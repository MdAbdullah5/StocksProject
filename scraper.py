import asyncio
import aiohttp
from bs4 import BeautifulSoup
from urllib.parse import urljoin
from io import BytesIO
from pdfminer.high_level import extract_text
from datetime import datetime


import re

def clean_excel_text(text):
    if not text:
        return ""
    # Remove illegal characters using openpyxl's recommended pattern
    ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
    return ILLEGAL_CHARACTERS_RE.sub("", text)



#-------------------Save to Excel------------------------
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

def save_to_excel(results, filename="company_reports.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report Summary"

    headers = [
        "SNo", "Company", "Annual Report URL", "Quarterly Report URL", "Press Release URL",
        "Annual Text", "Quarterly Text", "Press Text",
        "Start Time", "End Time", "Status"
    ]

    ws.append(headers)
    for col_num in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_num).font = Font(bold=True)

    for i, row in enumerate(results, start=1):
        ws.append([
            i,
            row.get("company_url", ""),
            row.get("annual_pdf", ""),
            row.get("quarterly_pdf", ""),
            row.get("press_release_url", ""),
            clean_excel_text(row.get("annual_text", "")[:500]),
            clean_excel_text(row.get("quarterly_text", "")[:500]),
            clean_excel_text(row.get("press_text", "")[:500]),
            row.get("start_time", ""),
            row.get("end_time", ""),
            row.get("status", "Failed")
        ])

    for i, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 30

    wb.save(filename)
    print(f"\u2705 Excel report saved to: {filename}")

# --------------------- CONFIG ---------------------
#-----the specific line sends requets like browser
headers = {'User-Agent': 'Mozilla/5.0'}

####### use to text the scraper

# urls = [
#     'https://www.029-group.com/investor-relations',
#     'https://www.1und1.ag/investor-relations-en',
#     'https://www.123fahrschule.de/news-and-reports',
#     'https://2invest-ag.com/finanzberichterstattung/',
#     'https://www.3u.net/en/investor-relations/publications/',
#     'https://www.4sc.com/investors/investor-information/financial-reports/'
# ]

# ------------------- HELPERS ----------------------
#-----------------the function used to send request and get content using requests module--------------
async def fetch(session, url):
    try:
        async with session.get(url, timeout=20) as response:
            if response.status == 200:
                return await response.read()
    except Exception as e:
        print(f"Error fetching {url}: {e}")
    return None


# here we get the urls using a tags which contain annual, quarterly, press as keywords
async def get_pdf_links_from_page(session, url):
    html = await fetch(session, url)
    if not html:
        return {}

    soup = BeautifulSoup(html, 'html.parser')
    links = soup.find_all('a', href=True)

    result = {
        "annual_pdf": None,
        "quarterly_pdf": None,
        "press_release_url": None,
    }

    seen = set()
    for tag in links:
        href = tag['href']
        full_url = urljoin(url, href)
        href_lower = href.lower()

        if result["annual_pdf"] is None and href_lower.endswith(".pdf") and "annual" in href_lower:
            result["annual_pdf"] = full_url
            seen.add("annual")

        elif result["quarterly_pdf"] is None and href_lower.endswith(".pdf") and "quarterly" in href_lower:
            result["quarterly_pdf"] = full_url
            seen.add("quarterly")

        elif result["press_release_url"] is None and any(k in href_lower for k in ["press", "news", "media", "announcement"]):
            result["press_release_url"] = full_url
            seen.add("press")

        if len(seen) == 3:
            break

    return result



######## here the pdf extraction is done ########
async def download_and_extract_pdf(session, pdf_url):
    pdf_bytes = await fetch(session, pdf_url)
    if not pdf_bytes:
        return None

    try:
        text = extract_text(BytesIO(pdf_bytes))
        return text.strip()
    except Exception as e:
        print(f"Error parsing PDF from {pdf_url}: {e}")
        return None


###### here all the functions are called and processed#########
async def process_company(session, url):
    result = {
        "company_url": url,
        "annual_pdf": None,
        "quarterly_pdf": None,
        "press_release_url": None,
        "annual_text": "",
        "quarterly_text": "",
        "press_text": "",
        "start_time": "",
        "end_time": "",
        "status": "Failed"
    }

    try:
        start = datetime.now()
        result["start_time"] = start.strftime('%Y-%m-%d %H:%M:%S')

        pdf_links = await get_pdf_links_from_page(session, url)

        if pdf_links.get("annual_pdf"):
            result["status"] = "Success"
            result["annual_pdf"] = pdf_links["annual_pdf"]
            result["annual_text"] = await download_and_extract_pdf(session, result["annual_pdf"]) or ""

        if pdf_links.get("quarterly_pdf"):
            result["status"] = "Success"
            result["quarterly_pdf"] = pdf_links["quarterly_pdf"]
            result["quarterly_text"] = await download_and_extract_pdf(session, result["quarterly_pdf"]) or ""

        if pdf_links.get("press_release_url"):
            result["status"] = "Success"
            result["press_release_url"] = pdf_links["press_release_url"]
            if pdf_links["press_release_url"].lower().endswith(".pdf"):
                result["press_text"] = await download_and_extract_pdf(session, result["press_release_url"]) or ""
            else:
                result["press_text"] = "Press release page (not PDF)"

        end = datetime.now()
        result["end_time"] = end.strftime('%Y-%m-%d %H:%M:%S')


    except Exception as e:
        result["status"] = f"Error: {str(e)}"

    return result

# ------------------- MAIN ----------------------

async def main(company_urls):
    async with aiohttp.ClientSession(headers=headers) as session:
        tasks = [process_company(session, url) for url in company_urls]
        all_results = await asyncio.gather(*tasks)

    save_to_excel(all_results)

